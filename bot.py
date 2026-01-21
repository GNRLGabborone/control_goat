import discord
from discord.ext import commands
import pandas as pd
import io
import os
from openpyxl.styles import PatternFill, Font, Alignment

intents = discord.Intents.default()
intents.members = True 
intents.message_content = True
bot = commands.Bot(command_prefix="!", intents=intents)

def limpiar(t):
    return t.strip().lower() if t else ""

@bot.command(name="reporte")
async def generar_reporte(ctx, message_id: int):
    try: await ctx.message.delete()
    except: pass

    NOMBRE_ROL_OBJETIVO = "GOAT" 

    # Forzar carga de miembros
    await ctx.guild.chunk()

    # Búsqueda Global
    msg = None
    for canal in ctx.guild.text_channels:
        try:
            msg = await canal.fetch_message(message_id)
            if msg: break
        except: continue

    if not msg or not msg.embeds:
        await ctx.send(f"❌ No encontré la convocatoria con ID `{message_id}`.")
        return

    embed = msg.embeds[0]
    titulo_convocatoria = embed.title if embed.title else "Convocatoria"
    
    lineas_si = ""
    lineas_no = ""
    for field in embed.fields:
        if "Accept" in field.name or "Aceptado" in field.name:
            lineas_si = limpiar(field.value)
        elif "Declin" in field.name or "Rechazado" in field.name:
            lineas_no = limpiar(field.value)

    lista_completa = []
    menciones_goat = []
    total_goat = 0
    anotados_si = 0
    anotados_no = 0
    faltan_goat = 0
    
    for miembro in ctx.guild.members:
        if miembro.bot: continue
        
        es_goat = any(rol.name == NOMBRE_ROL_OBJETIVO for rol in miembro.roles)
        n_user = limpiar(miembro.name)
        n_nick = limpiar(miembro.display_name)
        n_id = str(miembro.id)
        
        estado_texto = "FALTA"
        icono = "⚠️" # Triángulo amarillo para faltantes
        
        if n_user in lineas_si or n_nick in lineas_si or n_id in lineas_si:
            estado_texto = "SÍ"
            icono = "✅" # Tilde verde
        elif n_user in lineas_no or n_nick in lineas_no or n_id in lineas_no:
            estado_texto = "NO"
            icono = "❌" # Cruz roja
            
        if es_goat:
            total_goat += 1
            if estado_texto == "SÍ": anotados_si += 1
            elif estado_texto == "NO": anotados_no += 1
            else:
                faltan_goat += 1
                menciones_goat.append(miembro.mention)
        
        lista_completa.append({
            "Usuario": miembro.display_name, 
            "Estado": f"{icono} {estado_texto}", 
            "Es GOAT": "SÍ" if es_goat else "NO"
        })

    # --- EXCEL CON COLORES ACTUALIZADOS ---
    df = pd.DataFrame(lista_completa).sort_values(by="Estado")
    
    with io.BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Asistencia', index=False)
            ws = writer.sheets['Asistencia']
            
            # Definir colores
            v_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde
            r_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Rojo
            a_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarillo (para advertencia)
            
            for row in range(2, len(df) + 2):
                celda = ws.cell(row=row, column=2)
                valor = celda.value
                if "SÍ" in valor: celda.fill = v_fill
                elif "NO" in valor: celda.fill = r_fill
                elif "FALTA" in valor: celda.fill = a_fill
                celda.alignment = Alignment(horizontal="center")

            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12

        buffer.seek(0)
        nombre_f = "".join(x for x in titulo_convocatoria if x.isalnum() or x in "._- ").strip()
        await ctx.send(f"📋 **REPORTE DE ASISTENCIA**\n🔹 **Evento:** `{titulo_convocatoria}`", 
                       file=discord.File(fp=buffer, filename=f"Reporte_{nombre_f}.xlsx"))

    # Resumen Estadístico Detallado
    resumen = (
        f"📊 **Estadísticas @{NOMBRE_ROL_OBJETIVO}:**\n"
        f"✅ Anotados (SÍ): `{anotados_si}`\n"
        f"❌ Rechazados (NO): `{anotados_no}`\n"
        f"⚠️ Faltan anotar: `{faltan_goat}`\n"
        f"👥 **Total {NOMBRE_ROL_OBJETIVO}:** `{total_goat}`"
    )
    await ctx.send(resumen)

    if menciones_goat:
        header = f"⚠️ **Soldados @{NOMBRE_ROL_OBJETIVO} pendientes en '{titulo_convocatoria}':**\n"
        actual = header
        for m in menciones_goat:
            if len(actual) + len(m) > 1900:
                await ctx.send(actual)
                actual = ""
            actual += f"{m} "
        await ctx.send(actual)

TOKEN = os.environ.get('DISCORD_TOKEN') or 'MTQ2MzMwMDUzNTg2NzI4MTYwNg.GEatiw.BJUKA-T2Oi-3YJXHGsgBesgzwpvNKBMdWNBrRM'
bot.run(TOKEN)